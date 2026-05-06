from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Executive Summary"

# ── Styles ──────────────────────────────────────────────────────────────────
def side(style="thin", color="2E4057"):
    return Side(style=style, color=color)

thin_border = Border(left=side(), right=side(), top=side(), bottom=side())

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E4057"
SEC_BLUE    = "2F5496"
YELLOW_BG   = "FFF2CC"
GREEN_BG    = "E2EFDA"
ALT_BG      = "EEF2F7"
WHITE       = "FFFFFF"
NORMAL_BG   = "FFFFFF"

def apply(cell, value="", bold=False, size=10, color="000000", bg=None,
          wrap=True, halign="left", valign="top", border=True):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = thin_border

# ── Title block ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
apply(ws["A1"], "INTERNAL AUDIT EXECUTIVE SUMMARY", bold=True, size=16,
      color=WHITE, bg=DARK_BLUE, halign="center", valign="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
apply(ws["A2"], "S. Pyarelal Ispat Pvt. Ltd.  |  Period: January 2026 to March 2026",
      bold=True, size=11, color=WHITE, bg=MID_BLUE, halign="center", valign="center")
ws.row_dimensions[2].height = 20

ws.merge_cells("A3:F3")
apply(ws["A3"], "Internal Auditors: Agrawal Jain & Co., Chartered Accountants, Raipur",
      size=9, color=WHITE, bg=SEC_BLUE, halign="center", valign="center")
ws.row_dimensions[3].height = 16

# Legend row
ws.merge_cells("A4:F4")
apply(ws["A4"],
      "Legend:   🟡 Yellow = Financial Impact / Critical Observation     🟢 Green = No Reportable Discrepancy",
      size=9, color="444444", bg="F0F4FA", halign="center", valign="center")
ws.row_dimensions[4].height = 16

# ── Column headers ───────────────────────────────────────────────────────────
headers = ["S.No.", "Particulars", "Financial Impact", "Observation",
           "Annexure No.", "Suggestion"]
col_widths = [6, 30, 28, 45, 13, 42]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=5, column=col)
    apply(cell, h, bold=True, size=10, color=WHITE, bg=DARK_BLUE,
          halign="center", valign="center")
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[5].height = 22

# ── Data ─────────────────────────────────────────────────────────────────────
# flag: "critical"=yellow, "ok"=green, "normal"=white/alt
rows_data = [
    # ── STATUTORY DUES ────────────────────────────────────────────────────────
    ("SEC", "STATUTORY DUES & COMPLIANCES"),
    ("1","TDS Short Deduction (Sec 194Q, 194C, 194I, 194J, 194H)",
     "Short deduction of TDS (amount unquantified)",
     "Instances observed where TDS short deducted u/s 194Q; short deduction also under 194C, 194I, 194J, 194H.",
     "A1, A2",
     "Review all vendor transactions; recalculate correct TDS liability and deposit the shortfall. Strengthen internal controls on TDS applicability checks.",
     "critical"),
    ("2","TCS Short Collection (Sec 206C(1))",
     "Short collection of TCS (amount unquantified)",
     "Instances observed where TCS required to be collected u/s 206C(1) is short collected.",
     "A3",
     "Identify all transactions attracting TCS; collect pending TCS and deposit immediately. Set up automated alerts in ERP for TCS applicability.",
     "critical"),
    ("3","Delayed Deposit of TDS/TCS",
     "Interest paid: ₹15,352/-\nDisallowable under IT Act: ₹3,864/-",
     "Company incurred interest expenses of ₹15,352/- due to delayed deposit of TDS/TCS. ₹3,864/- is not deductible under the Income Tax Act.",
     "A4, A7",
     "Set up calendar-based reminders for TDS/TCS deposit due dates. Assign dedicated responsibility to treasury team for timely payment.",
     "critical"),
    ("4","TRACES Demand Outstanding",
     "Processed demand: ₹1,50,280/-",
     "A processed demand of ₹1,50,280/- raised on TRACES portal due to short payment, late payment, interest, or late filing fees under various TDS sections.",
     "A5",
     "Immediately review TRACES demand, reconcile with internal records and either deposit dues or file rectification request as applicable.",
     "critical"),
    ("5","Unconsumed Challans on TRACES",
     "No direct financial impact;\nRisk of lapsed challans",
     "40 challans remain unconsumed in TRACES — excess TDS deposited that has not been tagged to deductees.",
     "A6",
     "Map all 40 unconsumed challans to corresponding deductee entries. Initiate correction statements on TRACES to consume these challans.",
     "normal"),
    ("6","Audit Trail / Edit Log Non-Compliance\n(Companies (Accounts) Amendment Rules, 2021)",
     "Penalty: ₹50,000 (min) to ₹5,00,000 (max) under Sec 128",
     "High volume of back-dated entries: 3,489 purchase; 10,985 sales; 20,393 issue; 3,757 bank; 3,412 GRN (11,343 workflow edits). Significant deletions also noted.",
     "A8, A9",
     "Renew accounting software license with audit trail/edit log feature enabled. Restrict back-dating through system controls; require management approval for corrections.",
     "critical"),
    ("7","Income Tax Demand Outstanding",
     "Amount as per IT portal (unquantified)",
     "Demand found outstanding in Income Tax portal; regular follow-up not maintained adequately.",
     "A10",
     "Review all outstanding IT demands. Engage tax consultant to file responses/appeals within due dates to avoid further penalties.",
     "critical"),
    ("8","Employee Provident Fund (EPF) Discrepancies",
     "Potential liability for F.Y. 2025-26 (unquantified)",
     "Reportable discrepancies observed in adequacy of EPF deduction, contribution and payment for F.Y. 2025-26.",
     "A11",
     "Reconcile EPF workings with actual salaries. Deposit any shortfall with penal interest. Ensure monthly challan is deposited by due date.",
     "critical"),
    ("9","Employee State Insurance Corporation (ESIC) Discrepancies",
     "Potential liability for F.Y. 2025-26 (unquantified)",
     "Reportable discrepancies observed in adequacy of ESIC deduction, contribution and payment for F.Y. 2025-26.",
     "A11",
     "Reconcile ESIC workings, correct misclassification of employees and deposit pending dues. Review monthly compliance process.",
     "critical"),
    ("10","Non-Compliance of Sec 73 & 74 – Companies Act, 2013 (Deposits)",
     "Penalties applicable for non-compliance",
     "Discrepancies observed in compliance with Sec 73 and 74 of Companies Act, 2013 relating to acceptance of deposits.",
     "A12",
     "Ensure all deposits accepted in compliance with prescribed conditions. File required returns (DPT-3 etc.) with the Registrar of Companies immediately.",
     "critical"),
    ("11","CSR Expenditure Non-Compliance",
     "Unspent CSR amount to be transferred to specified funds",
     "Discrepancies observed in CSR expenditure; mandatory CSR spending may not have been fully met.",
     "A13",
     "Compute mandatory CSR obligation for the year. Transfer unspent amount (if any) to designated fund as per Companies Act provisions.",
     "critical"),
    ("12","Loss of Interest on CC due to Form 27C Not Obtained",
     "Loss of interest: ₹28,971/-",
     "Form 27C not obtained from parties, resulting in TCS being collected on scrap purchases, causing loss of interest on CC account of ₹28,971/-.",
     "A15",
     "Obtain Form 27C from eligible buyers to avoid TCS on scrap. Maintain a register to track 27C submissions and follow up proactively.",
     "critical"),
    ("13","Miscellaneous Observations",
     "Potential impact on P&L and Balance Sheet classification",
     "(1) Interest on CSPDCL for F.Y. 2025-26 not booked.\n(2) Loan Against Land classified under Unsecured Loan instead of Secured Loan.",
     "A16",
     "(1) Book CSPDCL interest in correct period for accurate P&L.\n(2) Reclassify Loan Against Land to Secured Loans to correctly reflect security structure.",
     "normal"),
    # ── GST ──────────────────────────────────────────────────────────────────
    ("SEC", "GOODS & SERVICE TAX (GST)"),
    ("14","GST Payable – Books vs GSTR-3B vs GSTR-1",
     "No financial impact",
     "GST payable as per books is in agreement with output liability shown in GSTR-1 and GSTR-3B. No reportable discrepancies.",
     "G1",
     "Continue the current reconciliation practice on a monthly basis.",
     "ok"),
    ("15","GST Under Reverse Charge Mechanism (RCM)",
     "Potential tax liability on unpaid RCM (unquantified)",
     "Discrepancies observed in GST payable under RCM — amount not properly paid as per applicable law.",
     "G2",
     "Identify all RCM-liable services/goods; discharge GST liability monthly and claim ITC in the same return period.",
     "critical"),
    ("16","RCM Liability Declared vs GSTR-2B Auto Population",
     "Mismatch leading to demand or excess payment risk",
     "Discrepancies observed between tax liability declared under RCM and amount auto-populated in GSTR-2B.",
     "G3",
     "Perform monthly reconciliation of self-assessed RCM against GSTR-2B auto-population. Resolve all differences before filing GSTR-3B.",
     "critical"),
    ("17","Blocked ITC Taken in Books (Sec 17(5))",
     "ITC reversal required (amount unquantified)",
     "Input Tax Credit taken in books on items blocked under Sec 17(5) of CGST Act.",
     "G4",
     "Identify and reverse all blocked ITC immediately. Set system-level check in ERP to flag Sec 17(5) transactions.",
     "critical"),
    ("18","ITC Reversal – Supplier Payment Beyond 180 Days",
     "ITC reversal required (amount unquantified)",
     "ITC not reversed where payment to suppliers has not been made within 180 days from the invoice date.",
     "G5",
     "Run ageing report of supplier invoices. Reverse ITC on all invoices beyond 180 days. Re-avail ITC once payment is made.",
     "critical"),
    ("19","GST TDS Pending Acceptance (March)",
     "Delay in availment of GST TDS credit",
     "GST TDS for the month of March has not been accepted on the GST portal, resulting in pending credit.",
     "G6",
     "Accept pending GST TDS credits on the GST portal immediately to ensure timely credit availment.",
     "normal"),
    ("20","GST TDS Credit Not Accounted",
     "Understated credit in books (unquantified)",
     "GST TDS credit accepted on portal has not been properly accounted in the books of accounts.",
     "G7",
     "Reconcile GST TDS credit available on portal with amounts in ERP. Account for all unrecorded credits to avoid understatement.",
     "normal"),
    ("21","GSTR-2B vs Books Mismatch",
     "Total mismatch: ₹9,29,661/-\n(ERP not in 2B: ₹2,28,666 | 2B not in ERP: ₹4,06,200 | C/D Note mismatch: ₹2,92,048 | Amount diff: ₹2,747)",
     "Invoices in ERP but not in GSTR-2B (C/D Note): ₹2,28,666; in GSTR-2B but not in ERP: ₹4,06,200; C/D Notes in 2B not in ERP: ₹2,92,048; Amount mismatch: ₹2,747.",
     "G8–G11",
     "Perform monthly GSTR-2B reconciliation before filing GSTR-3B. Follow up with vendors for missing invoices and update ERP for auto-populated invoices.",
     "critical"),
    ("22","ITC Claimed vs GSTR-2B Auto Population",
     "Risk of excess ITC claim or missed ITC",
     "Discrepancies observed between ITC claimed in GSTR-3B and ITC auto-populated in GSTR-2B.",
     "G12",
     "Reconcile ITC claim with GSTR-2B monthly. Restrict ITC claim to GSTR-2B eligible amount and reverse any excess claim.",
     "critical"),
    ("23","GRN Date Before Party Bill Date",
     "Risk of fictitious GRN / ITC availment",
     "Instances observed where GRN date is before party bill date — indicative of a process weakness.",
     "G13",
     "Introduce system-level validation in ERP to prevent GRN date from being earlier than invoice date. Review existing instances for genuineness.",
     "normal"),
    ("24","Purchases from Suppliers with Cancelled GST Registration",
     "ITC disallowance risk on such purchases",
     "Purchases made from suppliers whose GST registration has been cancelled.",
     "G14",
     "Verify supplier GSTIN status before raising purchase orders. Reverse ITC on such purchases and address with supplier.",
     "critical"),
    ("25","Suppliers with Delayed / Non-Filing of GSTR-3B",
     "ITC at risk due to supplier non-compliance",
     "Instances found where supplier GSTR-3B has not been filed or filed with significant delay, affecting ITC availability.",
     "G15",
     "Maintain a vendor compliance tracker. Issue notices to non-compliant vendors; consider withholding payments or switching vendors in chronic cases.",
     "critical"),
    ("26","GST Notices",
     "Potential liability/penalty depending on outcome",
     "GST notice received; reply has been uploaded. Ongoing monitoring required.",
     "G16",
     "Maintain a GST notice tracker. Ensure timely replies with supporting documents. Engage GST consultant for high-value notices.",
     "normal"),
    ("27","Filing of Monthly GST Returns (GSTR-1 & GSTR-3B)",
     "No financial impact",
     "GSTR-1 and GSTR-3B filed within statutory due dates for Jan, Feb and Mar 2026. No reportable discrepancies.",
     "—",
     "Continue timely filing. Set automated reminders for due dates to maintain compliance.",
     "ok"),
    # ── FINANCE ──────────────────────────────────────────────────────────────
    ("SEC", "FINANCE"),
    ("28","Excess Bank Interest Charged – PNB Cash Credit",
     "Excess interest charged: ₹1,17,176/-\n(Bank charged: ₹52,10,892 | As per workings: ₹50,93,716)",
     "Interest charged by PNB on CC account: ₹52,10,892/-; Interest as per internal calculation: ₹50,93,716/-. Excess charged: ₹1,17,176/-.",
     "B1",
     "Raise a formal dispute with PNB for the excess charge of ₹1,17,176/-. Conduct monthly reconciliation of bank interest with internal workings.",
     "critical"),
    ("29","Quarterly MSOD Return Submitted with Bank",
     "No direct financial impact; bank compliance risk",
     "Review of MSOD (Monthly Stock & Debtors) return submitted to bank conducted. Observations noted in annexure.",
     "MSOD",
     "Ensure MSOD is prepared accurately reflecting actual stock and debtors. Overstatement carries risk of recall of working capital limits.",
     "normal"),
    ("30","Bank Reconciliation Statement (BRS)",
     "No financial impact",
     "All bank current accounts and other accounts have been properly reconciled at end of the audit period. No reportable discrepancy.",
     "BRS",
     "Continue preparing BRS monthly. Unreconciled items should be investigated and cleared within 30 days.",
     "ok"),
    # ── AGEING ───────────────────────────────────────────────────────────────
    ("SEC", "AGEING ANALYSIS"),
    ("31","Sundry Creditors – Debit Balance (Advances to Suppliers)",
     "Risk of irrecoverable advances",
     "Advance payments made to suppliers for materials, expenses and transporters are outstanding; ageing analysis reveals long-outstanding items.",
     "AGE1",
     "Conduct reconciliation with all suppliers. Adjust advances against pending bills or recover amounts. Avoid fresh advances to suppliers with long-outstanding balances.",
     "normal"),
    ("32","Sundry Creditors – Credit Balance (Payables to Vendors)",
     "Risk of excess payment or unadjusted credits",
     "Payable balances to various creditors for materials, expenses and transporters contain long-outstanding items.",
     "AGE2",
     "Confirm balances with creditors through formal balance confirmation letters. Identify and adjust old credits no longer payable.",
     "normal"),
    ("33","Sundry Debtors – Debit Balance (Outstanding Receivables)",
     "Risk of bad debts / provision required",
     "Certain debtor debit balances are long-outstanding and require immediate attention and follow-up.",
     "AGE3",
     "Initiate collection drives for long-outstanding debtors. Consider making provision for doubtful debts as per accounting standards. Escalate chronic cases for legal action.",
     "normal"),
    ("34","Sundry Debtors – Credit Balance",
     "Risk of unadjusted excess receipts",
     "Long-outstanding credit balances in debtor accounts need attention — may represent advances received or excess payments.",
     "AGE4",
     "Reconcile with debtors. Adjust credit balances against pending invoices or refund where applicable.",
     "normal"),
    # ── POWER ────────────────────────────────────────────────────────────────
    ("SEC", "ANALYSIS OF POWER GENERATION & SALE"),
    ("35","Verification of Load Factor Incentive (Power Division)",
     "No financial impact",
     "Load factor incentive received from CSPDCL verified as per applicable tariff order. No discrepancies found.",
     "E1",
     "Continue monitoring load factor incentive quarterly to ensure correct claim as per tariff orders.",
     "ok"),
]

# ── Write rows ────────────────────────────────────────────────────────────────
current_row = 6
alt = False

for item in rows_data:
    if item[0] == "SEC":
        # Section header spanning all 6 columns
        ws.merge_cells(f"A{current_row}:F{current_row}")
        cell = ws[f"A{current_row}"]
        apply(cell, item[1], bold=True, size=10, color=WHITE, bg=SEC_BLUE,
              halign="center", valign="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        alt = False
        continue

    sno, particulars, impact, observation, annexure, suggestion, flag = item

    if flag == "critical":
        bg = YELLOW_BG
    elif flag == "ok":
        bg = GREEN_BG
    else:
        bg = ALT_BG if alt else NORMAL_BG
        alt = not alt

    row_vals = [sno, particulars, impact, observation, annexure, suggestion]
    for col, val in enumerate(row_vals, start=1):
        cell = ws.cell(row=current_row, column=col)
        halign = "center" if col in (1, 5) else "left"
        apply(cell, val, size=9, bg=bg, halign=halign, valign="top")

    # Auto height approximation
    max_lines = max(str(v).count("\n") + len(str(v)) // 60 + 1 for v in row_vals)
    ws.row_dimensions[current_row].height = max(30, min(max_lines * 13, 130))
    current_row += 1

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "A6"

# ── Summary counts sheet ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16

ws2.merge_cells("A1:B1")
apply(ws2["A1"], "AUDIT FINDINGS SUMMARY", bold=True, size=12, color=WHITE,
      bg=DARK_BLUE, halign="center", valign="center")
ws2.row_dimensions[1].height = 24

headers2 = [("Category", DARK_BLUE), ("No. of Findings", DARK_BLUE)]
for col, (h, bg) in enumerate(headers2, start=1):
    cell = ws2.cell(row=2, column=col)
    apply(cell, h, bold=True, size=10, color=WHITE, bg=bg, halign="center")
ws2.row_dimensions[2].height = 18

summary_data = [
    ("Statutory Dues & Compliances", 13),
    ("Goods & Service Tax (GST)", 14),
    ("Finance", 3),
    ("Ageing Analysis", 4),
    ("Power Generation & Sale", 1),
    ("TOTAL FINDINGS", 35),
]
for r, (label, count) in enumerate(summary_data, start=3):
    is_total = label.startswith("TOTAL")
    bg2 = DARK_BLUE if is_total else (ALT_BG if r % 2 == 0 else NORMAL_BG)
    col_color = WHITE if is_total else "000000"
    c1 = ws2.cell(row=r, column=1)
    c2 = ws2.cell(row=r, column=2)
    apply(c1, label, bold=is_total, size=10, color=col_color, bg=bg2)
    apply(c2, count, bold=is_total, size=10, color=col_color, bg=bg2, halign="center")
    ws2.row_dimensions[r].height = 18

ws2.row_dimensions[9].height = 22

# Critical vs OK
ws2.merge_cells("A10:B10")
apply(ws2["A10"], "FINDING STATUS BREAKDOWN", bold=True, size=10, color=WHITE,
      bg=MID_BLUE, halign="center", valign="center")
ws2.row_dimensions[10].height = 18

status_data = [
    ("🟡 Critical / Financial Impact Findings", 23, YELLOW_BG),
    ("🟢 No Discrepancy Found", 3, GREEN_BG),
    ("⚪ Observations / Process Gaps", 9, ALT_BG),
]
for r, (label, count, bg2) in enumerate(status_data, start=11):
    c1 = ws2.cell(row=r, column=1)
    c2 = ws2.cell(row=r, column=2)
    apply(c1, label, size=10, bg=bg2)
    apply(c2, count, size=10, bg=bg2, halign="center")
    ws2.row_dimensions[r].height = 18

# Key financial impacts
ws2.merge_cells("A15:B15")
apply(ws2["A15"], "KEY QUANTIFIED FINANCIAL IMPACTS", bold=True, size=10, color=WHITE,
      bg=MID_BLUE, halign="center", valign="center")
ws2.row_dimensions[15].height = 18

fin_data = [
    ("TRACES Demand Outstanding", "₹1,50,280"),
    ("Interest on Delayed TDS/TCS Deposit", "₹15,352"),
    ("Disallowable TDS Interest (IT Act)", "₹3,864"),
    ("Loss of Interest on CC (27C not obtained)", "₹28,971"),
    ("Excess Bank Interest Charged by PNB", "₹1,17,176"),
    ("GSTR-2B vs Books Mismatch (GST amount)", "₹9,29,661"),
]
for r, (label, amt) in enumerate(fin_data, start=16):
    c1 = ws2.cell(row=r, column=1)
    c2 = ws2.cell(row=r, column=2)
    apply(c1, label, size=9, bg=YELLOW_BG)
    apply(c2, amt, size=9, bg=YELLOW_BG, halign="right")
    ws2.row_dimensions[r].height = 16

wb.save("Executive_Summary_Pyarelal_Ispat_Jan_Mar_2026.xlsx")
print("Done")
